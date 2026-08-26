"""
analyze_data.py
================
Analyzes /mnt/session/outputs/parental_leave.csv (a copy of the original
/workspace/parental_leave.csv) and produces:

  1. /mnt/session/outputs/summary.json
     - total companies analyzed
     - average paid / unpaid maternity and paternity leave (weeks)
     - industry-level aggregations (mean, min, max) for each leave type
     - top 5 companies by total paid leave (paid maternity + paid paternity)

  2. /mnt/session/outputs/parental_leave_report.xlsx
     - 'Source Data'        : the raw dataset
     - 'Industry Summary'   : average leave by industry (formula-driven)
     - 'Leave Gap Analysis' : paid maternity vs paid paternity leave, by company
     - 'Summary'            : overall dataset statistics (formula-driven)

Notes / assumptions (also documented inside the workbook):
  - Source file is Latin-1 encoded (contains non-UTF8 bytes) and has four
    trailing empty "Unnamed" columns from stray trailing commas; both are
    handled on load.
  - Leave figures are expressed in weeks.
  - Blank cells in the source mean the figure was not reported for that
    company (NOT that the benefit is zero). Excel/statistics functions
    (AVERAGE, AVERAGEIF, MIN, MAX) ignore blanks automatically, so
    "average leave" figures are computed only over companies that reported
    that particular leave type.
  - For "Top 5 companies by total paid leave" and the Leave Gap Analysis,
    an unreported (blank) paid-leave figure is treated as 0 weeks so a
    total/gap can still be computed. This assumption is called out
    explicitly in the workbook and JSON output.
"""

import json
import os
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

SRC_CSV = "/mnt/session/outputs/parental_leave.csv"
JSON_OUT = "/mnt/session/outputs/summary.json"
XLSX_OUT = "/mnt/session/outputs/parental_leave_report.xlsx"

LEAVE_COLS = [
    "Paid Maternity Leave",
    "Unpaid Maternity Leave",
    "Paid Paternity Leave",
    "Unpaid Paternity Leave",
]

FONT_NAME = "Arial"


# --------------------------------------------------------------------------
# 1. Load & clean
# --------------------------------------------------------------------------
def load_data(path: str) -> pd.DataFrame:
    df = pd.read_csv(path, encoding="latin1")
    # Drop trailing empty "Unnamed" columns caused by stray trailing commas
    df = df.loc[:, ~df.columns.str.startswith("Unnamed")]
    # Trim whitespace on text columns
    df["Company"] = df["Company"].astype(str).str.strip()
    df["Industry"] = df["Industry"].fillna("Unknown").astype(str).str.strip()
    df.loc[df["Industry"].isin(["nan", "", "None"]), "Industry"] = "Unknown"
    # Ensure leave columns are numeric
    for col in LEAVE_COLS:
        df[col] = pd.to_numeric(df[col], errors="coerce")
    return df


# --------------------------------------------------------------------------
# 2. Build JSON summary
# --------------------------------------------------------------------------
def build_summary(df: pd.DataFrame) -> dict:
    total_companies = int(len(df))

    averages = {
        "Paid Maternity Leave (weeks)": round(float(df["Paid Maternity Leave"].mean()), 2),
        "Unpaid Maternity Leave (weeks)": round(float(df["Unpaid Maternity Leave"].mean()), 2),
        "Paid Paternity Leave (weeks)": round(float(df["Paid Paternity Leave"].mean()), 2),
        "Unpaid Paternity Leave (weeks)": round(float(df["Unpaid Paternity Leave"].mean()), 2),
    }

    industry_agg = {}
    grouped = df.groupby("Industry")
    for industry, g in grouped:
        entry = {"company_count": int(len(g))}
        for col in LEAVE_COLS:
            series = g[col].dropna()
            entry[col] = {
                "mean": round(float(series.mean()), 2) if len(series) else None,
                "min": round(float(series.min()), 2) if len(series) else None,
                "max": round(float(series.max()), 2) if len(series) else None,
                "n_reported": int(len(series)),
            }
        industry_agg[industry] = entry
    # sort industries alphabetically for stable output
    industry_agg = dict(sorted(industry_agg.items(), key=lambda kv: kv[0]))

    # Total paid leave = paid maternity + paid paternity; unreported (NaN) -> 0
    df_top = df.copy()
    df_top["Total Paid Leave (weeks)"] = (
        df_top["Paid Maternity Leave"].fillna(0) + df_top["Paid Paternity Leave"].fillna(0)
    )
    top5 = (
        df_top.sort_values("Total Paid Leave (weeks)", ascending=False)
        .head(5)[
            [
                "Company",
                "Industry",
                "Paid Maternity Leave",
                "Paid Paternity Leave",
                "Total Paid Leave (weeks)",
            ]
        ]
        .rename(
            columns={
                "Paid Maternity Leave": "Paid Maternity Leave (weeks)",
                "Paid Paternity Leave": "Paid Paternity Leave (weeks)",
            }
        )
    )
    top5_records = []
    for _, row in top5.iterrows():
        top5_records.append(
            {
                "Company": row["Company"],
                "Industry": row["Industry"],
                "Paid Maternity Leave (weeks)": None
                if pd.isna(row["Paid Maternity Leave (weeks)"])
                else round(float(row["Paid Maternity Leave (weeks)"]), 2),
                "Paid Paternity Leave (weeks)": None
                if pd.isna(row["Paid Paternity Leave (weeks)"])
                else round(float(row["Paid Paternity Leave (weeks)"]), 2),
                "Total Paid Leave (weeks)": round(float(row["Total Paid Leave (weeks)"]), 2),
            }
        )

    summary = {
        "assumptions": [
            "Leave figures are expressed in weeks.",
            "Blank/unreported values are excluded from averages, mins and maxes "
            "(they are not treated as zero).",
            "For 'Total Paid Leave' (used in the top-5 ranking) an unreported "
            "paid-leave figure is treated as 0 weeks so a total can be computed.",
            "3 companies (ASML, INK Communications Co., Rokt) had a blank "
            "Industry field and are grouped under 'Unknown'.",
        ],
        "total_companies_analyzed": total_companies,
        "average_leave_weeks": averages,
        "industry_aggregations": industry_agg,
        "top_5_companies_by_total_paid_leave": top5_records,
    }
    return summary


# --------------------------------------------------------------------------
# 3. Excel report helpers
# --------------------------------------------------------------------------
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=14, color="1F4E78")
NOTE_FONT = Font(name=FONT_NAME, italic=True, size=9, color="666666")
BODY_FONT = Font(name=FONT_NAME, size=10)
BOLD_FONT = Font(name=FONT_NAME, bold=True, size=10)
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header_row(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def autofit(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_source_data_sheet(wb, df: pd.DataFrame):
    ws = wb.create_sheet("Source Data")
    headers = ["Company", "Industry"] + LEAVE_COLS
    ws.append(headers)
    style_header_row(ws, 1, len(headers))
    for _, row in df.iterrows():
        vals = [row["Company"], row["Industry"]] + [
            None if pd.isna(row[c]) else float(row[c]) for c in LEAVE_COLS
        ]
        ws.append(vals)
    n = len(df)
    for r in range(2, n + 2):
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = BODY_FONT
            cell.border = BORDER
            if c >= 3:
                cell.number_format = "0.0"
                cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"
    autofit(ws, [32, 40, 14, 15, 14, 15])
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{n + 1}"
    return n, headers


def write_industry_summary_sheet(wb, df: pd.DataFrame, source_rows: int):
    ws = wb.create_sheet("Industry Summary")
    ws["A1"] = "Industry Summary — Average / Min / Max Leave by Industry (weeks)"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:N1")
    ws["A2"] = (
        "All figures pull live from the 'Source Data' sheet via AVERAGEIF / MINIFS / MAXIFS. "
        "Blank source cells are ignored automatically."
    )
    ws["A2"].font = NOTE_FONT
    ws.merge_cells("A2:N2")

    header_row = 4
    headers = ["Industry", "Company Count"]
    for col in LEAVE_COLS:
        headers += [f"{col} — Avg", f"{col} — Min", f"{col} — Max"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=c, value=h)
    style_header_row(ws, header_row, len(headers))

    industries = sorted(df["Industry"].unique())
    src_last = source_rows + 1  # last row number in Source Data (header is row1)

    for i, industry in enumerate(industries):
        r = header_row + 1 + i
        ws.cell(row=r, column=1, value=industry)
        ws.cell(row=r, column=2, value=f'=COUNTIF(\'Source Data\'!$B$2:$B${src_last},$A{r})')
        col_idx = 3
        for leave_col in LEAVE_COLS:
            src_col_letter = {
                "Paid Maternity Leave": "C",
                "Unpaid Maternity Leave": "D",
                "Paid Paternity Leave": "E",
                "Unpaid Paternity Leave": "F",
            }[leave_col]
            rng = f"'Source Data'!${src_col_letter}$2:${src_col_letter}${src_last}"
            crit_rng = "'Source Data'!$B$2:$B$" + str(src_last)
            # Gate on the count of *reported* (non-blank) values for this
            # industry/leave-type combo, so an industry with zero reported
            # values shows blank rather than a misleading 0 from MINIFS/MAXIFS.
            n_rep = f'COUNTIFS({crit_rng},$A{r},{rng},"<>")'
            ws.cell(
                row=r, column=col_idx,
                value=f'=IF({n_rep}=0,"",AVERAGEIF({crit_rng},$A{r},{rng}))',
            )
            ws.cell(
                row=r, column=col_idx + 1,
                value=f'=IF({n_rep}=0,"",_xlfn.MINIFS({rng},{crit_rng},$A{r}))',
            )
            ws.cell(
                row=r, column=col_idx + 2,
                value=f'=IF({n_rep}=0,"",_xlfn.MAXIFS({rng},{crit_rng},$A{r}))',
            )
            col_idx += 3

    last_data_row = header_row + len(industries)
    for r in range(header_row + 1, last_data_row + 1):
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = BODY_FONT
            cell.border = BORDER
            if c >= 2:
                cell.alignment = Alignment(horizontal="center")
            if c >= 3:
                cell.number_format = "0.0"

    ws.freeze_panes = f"A{header_row + 1}"
    autofit(ws, [40] + [16] * (len(headers) - 1))
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{last_data_row}"
    return last_data_row


def write_leave_gap_sheet(wb, df: pd.DataFrame, source_rows: int):
    ws = wb.create_sheet("Leave Gap Analysis")
    ws["A1"] = "Leave Gap Analysis — Paid Maternity vs Paid Paternity Leave by Company"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:F1")
    ws["A2"] = (
        "Gap (weeks) = Paid Maternity Leave − Paid Paternity Leave, pulled live from 'Source Data'. "
        "A blank paid-leave figure is treated as 0 weeks here so every company gets a gap value "
        "(see assumptions in the Summary sheet)."
    )
    ws["A2"].font = NOTE_FONT
    ws.merge_cells("A2:F2")

    header_row = 4
    headers = [
        "Company",
        "Industry",
        "Paid Maternity Leave (weeks)",
        "Paid Paternity Leave (weeks)",
        "Gap (Maternity − Paternity)",
        "Maternity Advantage?",
    ]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=c, value=h)
    style_header_row(ws, header_row, len(headers))

    src_last = source_rows + 1
    n = source_rows
    for i in range(n):
        r = header_row + 1 + i
        src_r = i + 2  # row in Source Data
        ws.cell(row=r, column=1, value=f"='Source Data'!A{src_r}")
        ws.cell(row=r, column=2, value=f"='Source Data'!B{src_r}")
        ws.cell(row=r, column=3, value=f"='Source Data'!C{src_r}")
        ws.cell(row=r, column=4, value=f"='Source Data'!E{src_r}")
        ws.cell(
            row=r, column=5,
            value=f"=IFERROR(N(C{r})-N(D{r}),\"\")",
        )
        ws.cell(
            row=r, column=6,
            value=f'=IF(E{r}>0,"Yes",IF(E{r}<0,"No (Paternity higher)","Equal"))',
        )

    last_data_row = header_row + n
    for r in range(header_row + 1, last_data_row + 1):
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = BODY_FONT
            cell.border = BORDER
            if c in (3, 4, 5):
                cell.number_format = "0.0"
                cell.alignment = Alignment(horizontal="center")
            if c == 6:
                cell.alignment = Alignment(horizontal="center")

    ws.freeze_panes = f"A{header_row + 1}"
    autofit(ws, [32, 40, 20, 20, 20, 22])
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{last_data_row}"
    return last_data_row


def write_summary_sheet(wb, source_rows: int, n_industries: int):
    ws = wb.create_sheet("Summary", 0)  # make it the first sheet
    ws["A1"] = "Parental Leave Dataset — Overall Summary"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:D1")

    src_last = source_rows + 1

    ws["A3"] = "Total companies analyzed"
    ws["B3"] = f"=COUNTA('Source Data'!$A$2:$A${src_last})"
    ws["A4"] = "Number of industries represented"
    ws["B4"] = n_industries

    stats_header_row = 6
    headers = ["Leave Type", "Average (weeks)", "Min (weeks)", "Max (weeks)", "Companies Reporting"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=stats_header_row, column=c, value=h)
    style_header_row(ws, stats_header_row, len(headers))

    col_letters = {
        "Paid Maternity Leave": "C",
        "Unpaid Maternity Leave": "D",
        "Paid Paternity Leave": "E",
        "Unpaid Paternity Leave": "F",
    }
    for i, leave_col in enumerate(LEAVE_COLS):
        r = stats_header_row + 1 + i
        col = col_letters[leave_col]
        rng = f"'Source Data'!${col}$2:${col}${src_last}"
        ws.cell(row=r, column=1, value=leave_col)
        ws.cell(row=r, column=2, value=f"=AVERAGE({rng})")
        ws.cell(row=r, column=3, value=f"=MIN({rng})")
        ws.cell(row=r, column=4, value=f"=MAX({rng})")
        ws.cell(row=r, column=5, value=f"=COUNT({rng})")

    last_stats_row = stats_header_row + len(LEAVE_COLS)
    for r in range(stats_header_row + 1, last_stats_row + 1):
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = BODY_FONT
            cell.border = BORDER
            if c >= 2:
                cell.alignment = Alignment(horizontal="center")
            if c in (2, 3, 4):
                cell.number_format = "0.0"

    ws["A3"].font = BOLD_FONT
    ws["A4"].font = BOLD_FONT
    ws["B3"].font = BODY_FONT
    ws["B4"].font = BODY_FONT

    notes_row = last_stats_row + 2
    ws.cell(row=notes_row, column=1, value="Assumptions & Notes:").font = BOLD_FONT
    notes = [
        "- Leave figures are expressed in weeks.",
        "- Averages/Min/Max are computed only over companies that reported that leave type; "
        "blank cells are ignored by AVERAGE/MIN/MAX/COUNT.",
        "- 3 companies (ASML, INK Communications Co., Rokt) had a blank Industry and are "
        "grouped as 'Unknown' in the Industry Summary sheet.",
        "- In the Leave Gap Analysis sheet, a blank paid-leave value is treated as 0 weeks "
        "so a gap can be computed for every company.",
        f"- Source: parental_leave.csv, {source_rows} company rows (originally read as Latin-1 "
        "encoded CSV with trailing empty columns removed).",
    ]
    for j, note in enumerate(notes, start=1):
        cell = ws.cell(row=notes_row + j, column=1, value=note)
        cell.font = NOTE_FONT
        ws.merge_cells(start_row=notes_row + j, start_column=1, end_row=notes_row + j, end_column=6)

    autofit(ws, [40, 20, 16, 16, 20])
    return ws


# --------------------------------------------------------------------------
# 4. Main
# --------------------------------------------------------------------------
def main():
    df = load_data(SRC_CSV)

    # ---- JSON summary ----
    summary = build_summary(df)
    with open(JSON_OUT, "w") as f:
        json.dump(summary, f, indent=2)
    print(f"Wrote {JSON_OUT}")

    # ---- Excel workbook ----
    wb = Workbook()
    # remove default sheet, we'll create named ones in desired order
    default_ws = wb.active
    wb.remove(default_ws)

    n_rows, _headers = write_source_data_sheet(wb, df)
    write_industry_summary_sheet(wb, df, n_rows)
    write_leave_gap_sheet(wb, df, n_rows)
    n_industries = df["Industry"].nunique()
    write_summary_sheet(wb, n_rows, n_industries)

    # Order sheets: Summary, Source Data, Industry Summary, Leave Gap Analysis
    order = ["Summary", "Source Data", "Industry Summary", "Leave Gap Analysis"]
    wb._sheets = [wb[name] for name in order]
    for name in order:
        wb[name].sheet_view.showGridLines = False

    wb.save(XLSX_OUT)
    print(f"Wrote {XLSX_OUT}")


if __name__ == "__main__":
    main()
